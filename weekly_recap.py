"""
Weekly Recap module — self-healing weekly digest built from the archive.

run_weekly_recap() (archive-only data builder):
1. Reads the archive for the last N days
2. Detects coverage gaps (companies with missing or stale data)
3. Backfills gaps by scraping + classifying (only what's missing)
4. Appends new signals to the archive
5. Deduplicates
6. Aggregates per-company signals and returns the data

The prepare/send orchestrators own the analysis, rendering, and email:
    python main.py --weekly-prepare   # build + draft key events, write JSON (no email)
    python main.py --weekly-send      # render saved JSON, email, publish
    python main.py --weekly-prepare --days 14   # custom lookback
"""

import json
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from config import (
    ANTHROPIC_API_KEY,
    ARCHIVE_PATH,
    BRIEF_OUTPUT_DIR,
    CATEGORIES,
    INDUSTRY_SECTORS,
    EMAIL_SENDER,
    EMAIL_PASSWORD,
    EMAIL_RECIPIENTS,
    SMTP_SERVER,
    SMTP_PORT,
)

_client = None

def _get_client():
    """Lazy-load Anthropic client only when synthesis is needed."""
    global _client
    if _client is None:
        from anthropic import Anthropic
        _client = Anthropic(api_key=ANTHROPIC_API_KEY)
    return _client


# ── Token & Cost Tracking ────────────────────────────────────────────

# Pricing per million tokens (Claude Sonnet 4.6)
MODEL_NAME = "claude-opus-4-8"
PRICE_INPUT_PER_M = 5.00    # $3.00 per 1M input tokens
PRICE_OUTPUT_PER_M = 25.00  # $15.00 per 1M output tokens

class TokenTracker:
    """Tracks cumulative token usage and cost across all API calls."""

    def __init__(self):
        self.calls = []
        self.total_input = 0
        self.total_output = 0

    def record(self, label: str, response):
        """Record usage from a Claude API response."""
        usage = getattr(response, "usage", None)
        if usage is None:
            return
        inp = getattr(usage, "input_tokens", 0)
        out = getattr(usage, "output_tokens", 0)
        cost = (inp / 1_000_000) * PRICE_INPUT_PER_M + (out / 1_000_000) * PRICE_OUTPUT_PER_M
        self.calls.append({"label": label, "input": inp, "output": out, "cost": cost})
        self.total_input += inp
        self.total_output += out

    @property
    def total_cost(self) -> float:
        return (self.total_input / 1_000_000) * PRICE_INPUT_PER_M + \
               (self.total_output / 1_000_000) * PRICE_OUTPUT_PER_M

    @property
    def call_count(self) -> int:
        return len(self.calls)

    def print_summary(self):
        """Print a formatted usage summary."""
        if not self.calls:
            print("  No API calls made")
            return

        print(f"\n  {'Label':<30} {'Input':>10} {'Output':>10} {'Cost':>10}")
        print(f"  {'─' * 62}")
        for c in self.calls:
            print(f"  {c['label']:<30} {c['input']:>10,} {c['output']:>10,} ${c['cost']:>8.4f}")
        print(f"  {'─' * 62}")
        print(f"  {'TOTAL':<30} {self.total_input:>10,} {self.total_output:>10,} ${self.total_cost:>8.4f}")
        print(f"\n  Model: {MODEL_NAME}")
        print(f"  Pricing: ${PRICE_INPUT_PER_M}/M input · ${PRICE_OUTPUT_PER_M}/M output")

# Module-level tracker — reset per run
_tracker = None

def _get_tracker() -> TokenTracker:
    global _tracker
    if _tracker is None:
        _tracker = TokenTracker()
    return _tracker

def _reset_tracker():
    global _tracker
    _tracker = TokenTracker()


# ── Step 1: Read last N days from archive ────────────────────────────

def read_archive_window(days: int = 7) -> dict:
    """
    Read the archive and return signals from the last N days,
    grouped by company.

    Returns:
        {
            "Stripe": [
                {"date": "2026-05-12", "category": "Fundraise / IPO", "headline": "...",
                 "summary": "...", "relevance": 5, "source": "Reuters", "url": "..."},
                ...
            ],
            "Anthropic": [...],
        }
    """
    path = Path(ARCHIVE_PATH)
    if not path.exists():
        print("[ERROR] No archive found. Run the daily scraper first.")
        return {}

    cutoff = (datetime.utcnow() - timedelta(days=days)).strftime("%Y-%m-%d")
    today = datetime.utcnow().strftime("%Y-%m-%d")

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    # Map header names to column indices
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False)):
        headers.append(str(cell.value or "").strip())

    col_map = {}
    expected = {
        "Date Scraped": "date",
        "Company": "company",
        "Category": "category",
        "Headline": "headline",
        "Summary": "summary",
        "Relevance": "relevance",
        "Source": "source",
        "URL": "url",
        "Published Date": "published_date",
        "Sector": "sector",
    }
    for i, h in enumerate(headers):
        if h in expected:
            col_map[expected[h]] = i

    by_company = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None for c in row):
            continue

        date_val = row[col_map.get("date", 0)]
        if date_val is None:
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val).strip()[:10]

        # Filter to window
        if date_str < cutoff or date_str > today:
            continue

        company = str(row[col_map.get("company", 1)] or "Unknown")
        category = str(row[col_map.get("category", 3)] or "Unknown")

        # Skip irrelevant items
        if category.lower() in ("irrelevant", "unknown"):
            continue

        item = {
            "date": date_str,
            "published_date": str(row[col_map.get("published_date", 8)] or date_str)[:10],
            "category": category,
            "headline": str(row[col_map.get("headline", 4)] or ""),
            "summary": str(row[col_map.get("summary", 5)] or ""),
            "relevance": int(row[col_map.get("relevance", 6)] or 1),
            "source": str(row[col_map.get("source", 7)] or "Unknown"),
            "url": str(row[col_map.get("url", 9)] or ""),
            "sector": str(row[col_map.get("sector", 2)] or ""),
        }

        by_company.setdefault(company, []).append(item)

    wb.close()

    # Sort each company's items by relevance (high first), then date (recent first)
    for company in by_company:
        by_company[company].sort(key=lambda x: (-x["relevance"], x["date"]), reverse=False)

    print(f"[INFO] Archive window: {cutoff} → {today}")
    print(f"[INFO] Found signals for {len(by_company)} companies")
    for name, items in sorted(by_company.items()):
        print(f"  {name}: {len(items)} signals")

    return by_company


def read_prior_signals(window_days: int = 7, lookback_days: int = 60) -> list[dict]:
    """Signals already archived in PRIOR weeks — older than the current window but
    within `lookback_days`. Used to drop this week's re-reports of old events
    (articles published this week about something covered weeks ago)."""
    recent = read_archive_window(days=lookback_days)
    window_start = (datetime.utcnow() - timedelta(days=window_days)).strftime("%Y-%m-%d")
    prior = []
    for company, items in recent.items():
        for it in items:
            d = (it.get("published_date") or it.get("date") or "")[:10]
            if d and d < window_start:
                p = dict(it)
                p["company"] = company
                prior.append(p)
    print(f"[INFO] Prior-week history: {len(prior)} signals older than {window_start} "
          f"(within {lookback_days}d)")
    return prior


# ── Step 2a: Local aggregation (no API calls) ────────────────────────

def aggregate_company(company_name: str, items: list[dict]) -> dict:
    """Build a recap dict purely from archive data — zero API cost."""

    # Derive top signal (highest relevance, most recent)
    sorted_items = sorted(items, key=lambda x: (-x["relevance"], x["published_date"]), reverse=False)
    top = sorted_items[0] if sorted_items else {}

    # Active categories
    cats = {}
    for item in items:
        cats[item["category"]] = cats.get(item["category"], 0) + 1
    active = sorted(cats.keys(), key=lambda c: -cats[c])

    # Simple auto-summary stats
    high_signals = [i for i in items if i["relevance"] >= 4]
    stat_line = (
        f"{len(items)} signal{'s' if len(items) != 1 else ''} detected across "
        f"{len(active)} categor{'ies' if len(active) != 1 else 'y'}. "
        f"{len(high_signals)} rated high-relevance (4-5 stars)."
    )
    if active:
        stat_line += f" Most active: {active[0]} ({cats[active[0]]} signal{'s' if cats[active[0]] != 1 else ''})."

    return {
        "company": company_name,
        "sector": items[0].get("sector", ""),
        "signal_count": len(items),
        "narrative": stat_line,
        "top_signal": top.get("summary", ""),
        "sentiment": "",  # No sentiment without synthesis
        "investor_takeaway": "",
        "active_categories": active,
        "signals": items,
        "synthesized": False,
    }


# ── Step 4: Send weekly email ────────────────────────────────────────

def send_weekly_email(html_body: str, week_label: str) -> bool:
    """Send the weekly recap email via SMTP."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    if not all([EMAIL_SENDER, EMAIL_PASSWORD, EMAIL_RECIPIENTS]):
        print("[WARN] Email credentials not configured. Skipping send.")
        return False

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"📊 Weekly Market Intel Recap — {week_label}"
    msg["From"] = EMAIL_SENDER
    msg["To"] = ", ".join(EMAIL_RECIPIENTS)

    plain = f"Weekly Market Intelligence Recap for {week_label}. View in an HTML-capable client."
    msg.attach(MIMEText(plain, "plain"))
    msg.attach(MIMEText(html_body, "html"))

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, EMAIL_RECIPIENTS, msg.as_string())
        print(f"[OK] Weekly recap sent to {', '.join(EMAIL_RECIPIENTS)}")
        return True
    except Exception as e:
        print(f"[ERROR] Weekly email send failed: {e}")
        return False


# ── Gap detection ────────────────────────────────────────────────────

def detect_coverage_gaps(by_company: dict, watchlist: list[dict], days: int) -> list[dict]:
    """
    Compare archive coverage against the watchlist.
    Returns a list of companies that need backfilling, each with a from_date.

    A company needs backfill if:
    - It has zero signals in the archive for the lookback window, OR
    - Its most recent signal is more than 1 day old (missed daily runs)
    """
    today = datetime.utcnow().strftime("%Y-%m-%d")
    cutoff = (datetime.utcnow() - timedelta(days=days)).strftime("%Y-%m-%d")
    yesterday = (datetime.utcnow() - timedelta(days=1)).strftime("%Y-%m-%d")

    gaps = []
    for company in watchlist:
        name = company["name"]
        items = by_company.get(name, [])

        if not items:
            # No coverage at all — full backfill from start of window
            gaps.append({"company": company, "from_date": cutoff, "reason": "no coverage"})
        else:
            # Find most recent signal date
            latest = max(item["date"] for item in items)
            if latest < yesterday:
                # Last signal is stale — backfill from that date
                gaps.append({"company": company, "from_date": latest, "reason": f"stale (last: {latest})"})

    return gaps


# ── Orchestrator ─────────────────────────────────────────────────────

def run_weekly_recap(days: int = 7, skip_analysis: bool = False, return_data: bool = False):
    """Archive-only weekly recap data builder with auto-backfill.

    Reads the archive, backfills coverage gaps, dedupes, and aggregates
    per-company signals from the archive (zero API cost). Always returns
    (recaps, None); the callers (run_weekly_prepare / run_weekly_send) own
    the analysis, rendering, and email steps.
    """
    from config import WATCHLIST
    from scraper import scrape_company
    from classifier import classify_articles
    from archiver import archive_items

    _reset_tracker()

    now = datetime.utcnow()
    start = now - timedelta(days=days)
    week_label = f"{start.strftime('%b %d')} – {now.strftime('%b %d, %Y')}"
    today = now.strftime("%Y-%m-%d")

    print(f"\n{'=' * 60}")
    print(f"  Weekly Market Intel Recap")
    print(f"  {week_label}")
    print(f"{'=' * 60}\n")

    # ── Step 1: Read current archive coverage ────────────────────
    print("Step 1: Checking archive coverage...")
    by_company = read_archive_window(days=days)

    # ── Step 2: Detect gaps ──────────────────────────────────────
    print(f"\nStep 2: Detecting coverage gaps...")
    gaps = detect_coverage_gaps(by_company, WATCHLIST, days=days)

    if gaps:
        print(f"  {len(gaps)} companies need backfill:")
        for g in gaps:
            print(f"    {g['company']['name']}: {g['reason']} → scraping from {g['from_date']}")

        # ── Step 3: Backfill gaps ────────────────────────────────
        print(f"\nStep 3: Backfilling {len(gaps)} companies...\n")
        backfill_signals = []

        for i, gap in enumerate(gaps, 1):
            company = gap["company"]
            name = company["name"]
            from_date = gap["from_date"]
            print(f"  [{i}/{len(gaps)}] {name} ({company.get('sector', '')})")
            print(f"    Scraping from {from_date} → {today}...")

            raw = scrape_company(company, from_date)
            print(f"    Found {len(raw)} articles")

            if raw:
                print(f"    Classifying...")
                classified = classify_articles(name, raw)
                relevant = [a for a in classified if a.get("category", "Irrelevant") != "Irrelevant"]
                backfill_signals.extend(relevant)
                print(f"    {len(relevant)} relevant signals")
            else:
                print(f"    No new articles found")

        # ── Step 4: Append backfill to archive ───────────────────
        if backfill_signals:
            print(f"\nStep 4: Archiving {len(backfill_signals)} backfilled signals...")
            archive_items(backfill_signals, WATCHLIST)
        else:
            print(f"\nStep 4: No new signals to archive")

        # ── Step 5: Re-read the now-complete archive ─────────────
        print(f"\nStep 5: Re-reading complete archive...")
        by_company = read_archive_window(days=days)
    else:
        print(f"  All {len(WATCHLIST)} companies have up-to-date coverage ✓")

    if not by_company:
        print("[WARN] No signals found in the archive for this period.")
        return [], None

    # ── Step 5.5: Deduplicate signals ────────────────────────────
    from deduplicator import deduplicate_by_company
    print(f"\nDeduplicating signals...")
    by_company = deduplicate_by_company(by_company, verbose=True)

    # ── Step 6: Aggregate per-company recaps from archive (0 API calls) ──
    print(f"\nStep 6: Aggregating from archive (0 API calls)...\n")
    recaps = []
    total = len(by_company)
    for i, (company, items) in enumerate(sorted(by_company.items()), 1):
        print(f"  [{i}/{total}] {company} ({len(items)} signals)")
        recap = aggregate_company(company, items)
        recaps.append(recap)

    # Summary
    backfill_count = len(gaps) if gaps else 0
    print(f"\n{'=' * 60}")
    print(f"  Weekly recap data ready")
    print(f"  Companies backfilled: {backfill_count}")
    print(f"  Companies in recap: {len(recaps)}")
    print(f"  Total signals: {sum(r['signal_count'] for r in recaps)}")
    print(f"{'=' * 60}\n")

    # Hand computed data to the prepare/send orchestrators.
    return recaps, None


#  Adds, using small targeted Claude calls (only for output the new
#  layout actually shows — no wasted per-company narratives):
#    • gather_private_movers()    – notable non-watchlist PRIVATE names (2nd segment)
#    • generate_connections()     – private↔public link per active watchlist company
#    • generate_week_in_brief()   – the editorial synthesis box
#    • draft_key_events()         – Saturday's draft of next-week events
#    • llm_dedupe()               – semantic dedupe pass (catches paraphrases)
#  Plus JSON persistence and the run_weekly_prepare() / run_weekly_send()
#  split for the prepare (Sat) → (optional edit) → send (Sun) flow.
#
#  NOTE: json, Path, datetime, timedelta are already imported at the top
#  of weekly_recap.py, so they are intentionally NOT re-imported here.
# ═══════════════════════════════════════════════════════════════════

WEEKLY_DIR = Path("weekly")
WEEKLY_DATA_PATH = WEEKLY_DIR / "weekly_data.json"
KEY_EVENTS_PATH = WEEKLY_DIR / "key_events_next_week.json"

# Optional: your GitHub Pages base for the footer "Browse the archive" link.
ARCHIVE_URL = ""   # e.g. "https://<your-user>.github.io/market-intel/"


def _chunks(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


def _parse_json(raw):
    """Tolerant JSON extraction: strips code fences, skips any leading prose,
    and decodes the FIRST complete JSON value — ignoring trailing 'extra data'
    (a common LLM habit that makes strict json.loads throw)."""
    raw = (raw or "").strip()
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[-1]
        if raw.rstrip().endswith("```"):
            raw = raw.rstrip()[:-3]
        raw = raw.strip()
    start = next((i for i, ch in enumerate(raw) if ch in "{["), None)
    if start is None:
        return json.loads(raw)  # no container found → raise a clear error
    obj, _ = json.JSONDecoder().raw_decode(raw[start:])
    return obj


def _sig_key(s):
    """Stable identity for a signal: real URL if present, else headline."""
    u = (s.get("url") or "").strip().lower()
    if u and u not in ("#", "n/a", "none"):
        return u
    return (s.get("headline") or "").strip().lower()


# ── Private↔Public connection note (batched, per active company) ─────
# A public company is surfaced ONLY when this week's signals show the watchlist
# (private) company directly dealing with it — never as a mere peer/read-through.
CONNECTION_PROMPT = """You are a private-markets analyst. Each entry below is a WATCHLIST \
(private) company and its notable signals THIS WEEK. Write the private->public connection in ONE \
sentence, but ONLY when this week's signals show the private company directly dealing with a \
specific public company — a partnership, contract, customer/supplier relationship, investment, \
acquisition (either direction), or the private company's own IPO / going-public step. Name that \
public company and state the dealing.
Hard rules:
- Pull in a public name ONLY because THIS watchlist company is dealing with it in this week's \
signals. NEVER mention a public company as a same-sector comparable, peer, or "read-through" — \
if the private company is not actually dealing with it this week, there is NO connection.
- If there is no such direct dealing this week, return an empty string for that company. Do not \
force or invent a connection.
Respond ONLY with valid JSON: {"connections": {"CompanyName": "one sentence or empty", ...}}."""


def generate_connections(recaps, chunk_size=10):
    """Draft a one-line private->public connection per active WATCHLIST company — only
    when that watchlist company is directly dealing with a public name this week. Writes
    onto recap['connection']; companies with no direct dealing are left without a note.
    Non-watchlist private movers are skipped (no connection note)."""
    active = [r for r in recaps if r.get("signal_count", 0) > 0 and r.get("watchlist", True)]
    result = {}
    for group in _chunks(active, chunk_size):
        blocks = []
        for r in group:
            sigs = sorted(r.get("signals", []),
                          key=lambda x: -int(x.get("relevance", 0) or 0))[:4]
            bullets = "\n".join(f"  - ({s.get('category')}) {s.get('headline', '')}" for s in sigs)
            blocks.append(f"{r.get('company')} [{r.get('sector', '')}]\n{bullets}")
        try:
            resp = _get_client().messages.create(
                model=MODEL_NAME, max_tokens=1200, system=CONNECTION_PROMPT,
                messages=[{"role": "user", "content": "\n\n".join(blocks)}],
            )
            _get_tracker().record("Connections", resp)
            result.update(_parse_json(resp.content[0].text).get("connections", {}))
        except Exception as e:
            print(f"  [ERROR] connection generation failed for a group: {e}")
    for r in recaps:
        note = (result.get(r.get("company")) or "").strip()
        if note:
            r["connection"] = note


# ── Week in brief (cheap synthesis from the week's top headlines) ────
BRIEF_PROMPT = """You are writing "The week in brief" for an investment desk — a 2-4 sentence \
editorial synthesis of the single most important throughline in private markets this week. Lead \
with the dominant theme; connect it to the private-public link; end on the one thing to watch. \
Prose only, no lists, no headline restatement.
Respond ONLY with valid JSON: {"week_in_brief": "..."}"""


def generate_week_in_brief(industry_news, top_signals):
    """Synthesise a short editorial brief from the week's themes (if any) + top headlines."""
    industry_news = industry_news or {}
    ctx = []
    themes = industry_news.get("themes", [])
    if themes:
        ctx.append("Themes:")
        for t in themes[:6]:
            ctx.append(f"  - {t.get('headline', '')}: {t.get('narrative', '')}")
    ctx.append("\nTop headlines this week:")
    for s in top_signals[:8]:
        ctx.append(f"  - {s.get('headline', '')}")
    if industry_news.get("market_sentiment"):
        ctx.append(f"\nMarket read: {industry_news['market_sentiment']}")
    try:
        resp = _get_client().messages.create(
            model=MODEL_NAME, max_tokens=400, system=BRIEF_PROMPT,
            messages=[{"role": "user", "content": "\n".join(ctx)}],
        )
        _get_tracker().record("Week in brief", resp)
        return _parse_json(resp.content[0].text).get("week_in_brief", "").strip()
    except Exception as e:
        print(f"  [ERROR] week-in-brief generation failed: {e}")
        return industry_news.get("market_sentiment", "")


# ── Key events draft ─────────────────────────────────────────────────
KEY_EVENTS_PROMPT = """You are a private-markets analyst drafting "Key Events - Next Week" for an \
investor newsletter, based on THIS WEEK's signals. Propose 3-5 forward-looking items worth \
watching next week: expected/rumored round closings, public-company earnings whose read-through \
affects these private names, IPO lockup expiries, notable conferences, or pending regulatory \
decisions.
Rules: include only events you can reasonably anticipate from the signals or well-known \
schedules. NEVER invent a precise calendar date you cannot support — if timing is uncertain use \
"Next week" or "Mid-week" for `when`. Mark exactly ONE item (the most important) with "key": true.
Respond ONLY with valid JSON: {"events": [{"when": "...", "label": "short bold lead-in", "text": \
"one sentence on what to watch and why", "key": true}]}"""


def draft_key_events(recaps, industry_news, days=7):
    """Claude drafts next-week key events from the week's signals. Returns a list."""
    from weekly_layout import flatten_signals, rank_and_split
    top, _ = rank_and_split(flatten_signals(recaps), top_n=15, radar_n=0)
    lines = [f"- ({s.get('category')}) {s.get('headline', '')}" for s in top]
    if industry_news and industry_news.get("market_sentiment"):
        lines.append(f"\nMarket read: {industry_news['market_sentiment']}")
    try:
        resp = _get_client().messages.create(
            model=MODEL_NAME, max_tokens=900, system=KEY_EVENTS_PROMPT,
            messages=[{"role": "user", "content": "\n".join(lines)}],
        )
        _get_tracker().record("Key events draft", resp)
        return _parse_json(resp.content[0].text).get("events", [])
    except Exception as e:
        print(f"  [ERROR] key-events draft failed: {e}")
        return []


# ── Semantic dedupe pass (catches reworded duplicates Jaccard misses) ─
SEMANTIC_DEDUP_PROMPT = """You are deduplicating a ranked list of news signals for an investor \
briefing. Some items are the SAME underlying story, reported by different outlets or reworded — \
even when the wording differs substantially. Group together items that describe the same \
underlying event or announcement about the same company or the same deal.
Strict rules:
- Only group items that are genuinely the same story. Different developments about the same \
company are NOT duplicates. Similar-sounding but distinct deals are NOT duplicates.
- Use the 0-based [n] indices shown.
Respond ONLY with valid JSON: {"duplicate_groups": [[i, j], [k, l, m], ...]}. Include only \
groups with 2 or more items; omit anything unique."""


def llm_dedupe(recaps, candidate_n=25):
    """Semantic dedupe over the top-ranked candidates. For each duplicate cluster,
    keeps the best-ranked item and removes the rest from `recaps` in place. Runs
    after Jaccard dedupe, so it only has to catch the reworded stragglers."""
    from weekly_layout import flatten_signals, rank_and_split

    top, _ = rank_and_split(flatten_signals(recaps), top_n=candidate_n, radar_n=0)
    if len(top) < 2:
        return

    lines = [f"[{i}] ({s.get('company', '?')} | {s.get('category', '')}) "
             f"{s.get('headline', '')} — {(s.get('summary') or '')[:160]}"
             for i, s in enumerate(top)]
    try:
        resp = _get_client().messages.create(
            model=MODEL_NAME, max_tokens=800, system=SEMANTIC_DEDUP_PROMPT,
            messages=[{"role": "user", "content": "\n".join(lines)}],
        )
        _get_tracker().record("Semantic dedupe", resp)
        groups = _parse_json(resp.content[0].text).get("duplicate_groups", [])
    except Exception as e:
        print(f"  [ERROR] LLM dedupe failed: {e}")
        return

    drop_pairs = set()  # (company, sig_key) of the losers in each cluster
    for group in groups:
        idxs = sorted(i for i in group if isinstance(i, int) and 0 <= i < len(top))
        for idx in idxs[1:]:  # keep the first (best-ranked); drop the rest
            s = top[idx]
            drop_pairs.add((s.get("company"), _sig_key(s)))

    if not drop_pairs:
        print("  LLM dedupe: no additional duplicates found")
        return

    removed = 0
    for r in recaps:
        kept = []
        for s in r.get("signals", []):
            if (r.get("company"), _sig_key(s)) in drop_pairs:
                removed += 1
                continue
            kept.append(s)
        r["signals"] = kept
        r["signal_count"] = len(kept)
    print(f"  LLM dedupe removed {removed} paraphrased duplicate(s)")


# ── Per-company semantic dedupe (guarantees every busy company is checked) ────
PER_COMPANY_DEDUP_PROMPT = """You are deduplicating this week's news signals company by company \
for an investor briefing. Under each COMPANY heading is a numbered list of that company's signals. \
Within EACH company, group the signals that report the SAME underlying event — the same funding \
round, deal, product launch, lawsuit, share sale, contract, or milestone — even if reworded, from \
different outlets, or written in a different language. Different developments about the same \
company (a separate round, an unrelated lawsuit, a different product) are NOT duplicates.
Match on the EVENT, not the wording. For example, "staff sold stock at a 10% discount", \
"employees cash in shares worth $85m", and "landmark $85m employee share sale" are all the SAME \
share-sale event and belong in one group. Same for a deal described once by the buyer's angle and \
once by the target's, or a figure quoted as "$39B" in one and "39 billion" in another.
Respond ONLY with valid JSON mapping each company to its groups of 2+ duplicate indices, e.g.:
{"Wayve": [[0, 1, 2]], "Kraken": [[1, 3]]}. Omit companies that have no duplicates."""


def llm_dedupe_per_company(recaps, min_signals=2, batch_companies=12):
    """For each company with >= min_signals signals, ask Claude to group same-event
    stories (reworded / cross-language / different outlets) and drop all but the
    best-ranked in each group. Batched across companies to limit calls. Unlike the
    global top-N pass, this guarantees every busy company is checked."""
    from weekly_layout import _rank_key

    targets = [r for r in recaps if len(r.get("signals", [])) >= min_signals]
    if not targets:
        return
    removed = 0
    for batch in _chunks(targets, batch_companies):
        ordered = {r.get("company"): list(r.get("signals", [])) for r in batch}
        blocks = []
        for r in batch:
            name = r.get("company")
            lines = [f"[{i}] {s.get('headline', '')} — {(s.get('summary') or '')[:120]}"
                     for i, s in enumerate(ordered[name])]
            blocks.append(f"COMPANY: {name}\n" + "\n".join(lines))
        try:
            resp = _get_client().messages.create(
                model=MODEL_NAME, max_tokens=1200, system=PER_COMPANY_DEDUP_PROMPT,
                messages=[{"role": "user", "content": "\n\n".join(blocks)}],
            )
            _get_tracker().record("Per-company dedupe", resp)
            groups_by_co = _parse_json(resp.content[0].text)
        except Exception as e:
            print(f"  [ERROR] per-company dedupe failed: {e}")
            continue
        if not isinstance(groups_by_co, dict):
            continue
        for r in batch:
            name = r.get("company")
            sigs = ordered.get(name, [])
            drop_idx = set()
            for group in (groups_by_co.get(name) or []):
                idxs = [i for i in group if isinstance(i, int) and 0 <= i < len(sigs)]
                if len(idxs) < 2:
                    continue
                # keep the best-ranked representative, drop the rest
                for i in sorted(idxs, key=lambda j: _rank_key(sigs[j]))[1:]:
                    drop_idx.add(i)
            if drop_idx:
                kept = [s for i, s in enumerate(sigs) if i not in drop_idx]
                removed += len(sigs) - len(kept)
                r["signals"] = kept
                r["signal_count"] = len(kept)
    print(f"  Per-company semantic dedupe removed {removed} duplicate(s)")


# ── Cross-time dedupe: drop this week's re-reports of already-covered events ──
RECENCY_PROMPT = """You are curating a weekly private-markets briefing. Each item pairs a NEW \
story (surfaced this week) with the MOST SIMILAR EARLIER story already covered in a prior week. \
For each, decide whether the NEW item reports a GENUINELY NEW development this week, or is just \
re-coverage / a retrospective of the SAME older event already reported.
Rules:
- "Re-report" = the same underlying event as the earlier story (same funding round, same deal, \
same launch, same lawsuit), merely re-published, rounded up, or reworded later. Mark these to DROP.
- A genuinely NEW development (a later or next round, a new milestone, a new contract, an \
escalation or new ruling) is NOT a re-report even if it resembles the earlier one. KEEP these.
- When genuinely unsure, KEEP (do not drop).
Respond ONLY with valid JSON: {"drop": [<indices of NEW items that are re-reports of the earlier story>]}."""


def _llm_recency_judge(pairs, max_pairs=30):
    """For each (this-week signal, most-similar earlier story) pair, ask Claude whether
    the new item is a re-report of the older event. Marks re-reports with s['_drop']=True.
    Returns the number dropped."""
    pairs = pairs[:max_pairs]
    if not pairs:
        return 0
    lines = []
    for i, (_r, s, p) in enumerate(pairs):
        lines.append(
            f"[{i}] NEW ({s.get('published_date', '')}): {s.get('headline', '')} — "
            f"{(s.get('summary') or '')[:160]}\n"
            f"    EARLIER ({p.get('published_date') or p.get('date', '')}): {p.get('headline', '')} — "
            f"{(p.get('summary') or '')[:160]}"
        )
    try:
        resp = _get_client().messages.create(
            model=MODEL_NAME, max_tokens=600, system=RECENCY_PROMPT,
            messages=[{"role": "user", "content": "\n\n".join(lines)}],
        )
        _get_tracker().record("Recency judge", resp)
        parsed = _parse_json(resp.content[0].text)
        # Accept {"drop": [...]} or a bare [...] array of indices
        if isinstance(parsed, dict):
            drop = parsed.get("drop", [])
        elif isinstance(parsed, list):
            drop = parsed
        else:
            drop = []
    except Exception as e:
        print(f"  [ERROR] recency judge failed: {e}")
        return 0
    n = 0
    for idx in drop:
        if isinstance(idx, int) and 0 <= idx < len(pairs):
            pairs[idx][1]["_drop"] = True
            n += 1
    return n


def filter_already_covered(recaps, prior, hard=0.40, soft=0.20):
    """Cross-time dedupe. Drop this week's signals that re-report an event already in
    the archive from a prior week. Deterministic hard-drop at Jaccard >= `hard`;
    borderline matches (>= `soft`) go to the LLM recency judge. Mutates recaps."""
    from weekly_layout import _tokenize, _jaccard, _same_subject

    prior_idx = []
    for p in prior:
        toks = _tokenize((p.get("headline") or "") + " " + (p.get("summary") or ""))
        if toks:
            prior_idx.append((toks, p))
    if not prior_idx:
        print("  No prior-week archive history to compare against — skipping")
        return

    hard_drops, borderline = 0, []
    for r in recaps:
        for s in r.get("signals", []):
            toks = _tokenize((s.get("headline") or "") + " " + (s.get("summary") or ""))
            if not toks:
                continue
            best, best_p = 0.0, None
            for ptoks, p in prior_idx:
                if not _same_subject(s, p):
                    continue
                j = _jaccard(toks, ptoks)
                if j > best:
                    best, best_p = j, p
            if best >= hard:
                s["_drop"] = True
                hard_drops += 1
            elif best >= soft and best_p is not None:
                borderline.append((r, s, best_p))

    llm_drops = _llm_recency_judge(borderline)

    removed = 0
    for r in recaps:
        kept = [s for s in r.get("signals", []) if not s.get("_drop")]
        removed += len(r.get("signals", [])) - len(kept)
        r["signals"] = kept
        r["signal_count"] = len(kept)
    print(f"  Already-covered filter: dropped {removed} re-reported signal(s) "
          f"({hard_drops} exact, {llm_drops} via recency judge)")


# ── Non-watchlist PRIVATE movers (broad scrape + significance + private gate) ─
# Broad, event-biased seed queries per sector (2 each = 12 total). Movers scrape
# RSS-only (use_newsapi=False in gather_private_movers) so it never touches the
# limited NewsAPI daily quota. Tune freely.
SECTOR_QUERIES = {
    "AI & Machine Learning": ["AI startup mega funding round", "AI lab valuation billion"],
    "Defense Tech & Aerospace": ["defense tech startup funding", "space startup funding round"],
    "Robotics & Hardware": ["robotics startup funding round", "humanoid robot startup funding"],
    "Fintech & Crypto": ["fintech startup funding billion", "crypto startup funding round"],
    "Energy & Nuclear": ["fusion energy startup funding", "nuclear startup funding"],
    "Social & Gaming": ["gaming startup funding round", "social app startup funding billion"],
}

PRIVATE_MOVERS_PROMPT = """You are curating a "Private Markets - Also Notable" section of a \
private-markets investor briefing. From the numbered headlines below, select ONLY genuinely \
notable developments about PRIVATE companies that are NOT on the watchlist.
Strict rules:
- The company MUST be privately held (venture/PE-backed or otherwise not publicly listed). \
EXCLUDE any publicly traded company and any subsidiary of a public company.
- EXCLUDE watchlist companies: {watchlist}
- The bar is high: mega funding rounds, major M&A involving a private company, notable new \
valuations, landmark contracts, or structurally significant events. Exclude routine product \
news, opinion/analysis, thin rumors, and anything primarily about a public company.
For each qualifying item, identify the primary PRIVATE company and explain why it is significant.
Respond ONLY with valid JSON:
{{"movers": [{{"index": <int, the [n] of the source headline>, "company": "Primary private company", \
"sector": "<one of: {sectors}>", "category": "<one allowed category>", "summary": \
"one-sentence factual summary", "why": "one sentence on why it's notable", "relevance": 5}}]}}
Allowed categories: {categories}
Return an empty list if nothing clears the bar. Quality over quantity - at most {max_movers}."""


def gather_private_movers(days=7, max_movers=8, max_articles=120):
    """Broad-scrape sector news, then have Claude keep only notable, NON-watchlist,
    PRIVATE companies (public names excluded). Returns recap-shaped dicts tagged
    watchlist=False, ready to merge into the recaps list as the second segment."""
    from config import WATCHLIST, INDUSTRY_SECTORS, CATEGORIES
    from scraper import scrape_company

    from_date = (datetime.utcnow() - timedelta(days=days)).strftime("%Y-%m-%d")
    wl_names = {w["name"].lower() for w in WATCHLIST}
    wl_alias = {a.lower() for w in WATCHLIST for a in w.get("aliases", [])}

    articles, seen = [], set()
    for sector, queries in SECTOR_QUERIES.items():
        for q in queries:
            try:
                # RSS-only: the movers scrape must not consume the NewsAPI quota
                raw = scrape_company({"name": q, "aliases": [], "sector": sector},
                                     from_date, use_newsapi=False)
            except Exception as e:
                print(f"  [WARN] private movers scrape failed for '{q}': {e}")
                continue
            for a in raw:
                k = (a.get("url") or "").split("?")[0].rstrip("/").lower()
                if not k or k in seen:
                    continue
                seen.add(k)
                a["_sector_guess"] = sector
                articles.append(a)

    if not articles:
        print("  No industry articles found for private movers")
        return []
    articles = articles[:max_articles]

    lines = [f"[{i}] ({a.get('source', '?')}, {a.get('published_date', '')}, "
             f"{a.get('_sector_guess')}) {a.get('title', '')} — {(a.get('snippet') or '')[:160]}"
             for i, a in enumerate(articles)]
    prompt = PRIVATE_MOVERS_PROMPT.format(
        watchlist=", ".join(sorted(w["name"] for w in WATCHLIST)),
        sectors=" | ".join(INDUSTRY_SECTORS),
        categories="; ".join(CATEGORIES),
        max_movers=max_movers,
    )
    try:
        resp = _get_client().messages.create(
            model=MODEL_NAME, max_tokens=1500, system=prompt,
            messages=[{"role": "user", "content": "\n".join(lines)}],
        )
        _get_tracker().record("Private movers", resp)
        movers = _parse_json(resp.content[0].text).get("movers", [])
    except Exception as e:
        print(f"  [ERROR] private movers LLM call failed: {e}")
        return []

    valid_cats = set(CATEGORIES)
    by_company = {}
    for m in movers[:max_movers]:
        idx = m.get("index")
        if not isinstance(idx, int) or not (0 <= idx < len(articles)):
            continue
        company = (m.get("company") or "").strip()
        if not company or company.lower() in wl_names or company.lower() in wl_alias:
            continue  # safety net: never re-include a watchlist name
        src = articles[idx]
        cat = m.get("category") if m.get("category") in valid_cats else "Company Timeline"
        sector = m.get("sector") or src.get("_sector_guess") or ""
        signal = {
            "category": cat,
            "sector": sector,
            "headline": src.get("title") or "",
            "summary": (m.get("summary") or "").strip(),
            "why": (m.get("why") or "").strip(),
            "relevance": int(m.get("relevance", 5) or 5),
            "source": src.get("source", "Unknown"),
            "published_date": src.get("published_date", ""),
            "url": src.get("url", "#"),
            "company": company,
        }
        rec = by_company.setdefault(company, {
            "company": company, "sector": sector, "signals": [],
            "signal_count": 0, "connection": "", "watchlist": False,
        })
        rec["signals"].append(signal)

    out = []
    for company, rec in by_company.items():
        rec["signal_count"] = len(rec["signals"])
        out.append(rec)
    print(f"  Private movers: {len(out)} non-watchlist private companies "
          f"({sum(r['signal_count'] for r in out)} signals)")
    return out


# ── JSON persistence ─────────────────────────────────────────────────
def _write_weekly_json(data, key_events):
    WEEKLY_DIR.mkdir(parents=True, exist_ok=True)
    with open(WEEKLY_DATA_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    with open(KEY_EVENTS_PATH, "w", encoding="utf-8") as f:
        json.dump(key_events, f, indent=2, ensure_ascii=False)
    print(f"[OK] Wrote {WEEKLY_DATA_PATH} and {KEY_EVENTS_PATH}")


def _read_weekly_json():
    if not WEEKLY_DATA_PATH.exists():
        return None, []
    with open(WEEKLY_DATA_PATH, encoding="utf-8") as f:
        data = json.load(f)
    key_events = []
    if KEY_EVENTS_PATH.exists():
        with open(KEY_EVENTS_PATH, encoding="utf-8") as f:
            key_events = json.load(f)
    return data, key_events


def _publish_html(html):
    """Write the rendered recap to GitHub Pages (docs/latest.html) + a dated copy."""
    docs = Path("docs")
    docs.mkdir(parents=True, exist_ok=True)
    (docs / "latest.html").write_text(html, encoding="utf-8")
    briefs = Path(BRIEF_OUTPUT_DIR)
    briefs.mkdir(parents=True, exist_ok=True)
    stamp = datetime.utcnow().strftime("%Y-%m-%d")
    (briefs / f"weekly_recap_{stamp}.html").write_text(html, encoding="utf-8")
    print(f"[OK] Published docs/latest.html and data/briefs/weekly_recap_{stamp}.html")


# ── Orchestrators: prepare (Saturday) / send (Sunday) ────────────────
def run_weekly_prepare(days: int = 7, issue: str = ""):
    """SATURDAY: build the recap (watchlist + notable non-watchlist private names),
    dedupe, draft key events, write JSON. No email."""
    print("\n=== WEEKLY PREPARE (draft) ===")
    recaps, industry_news = run_weekly_recap(days=days, skip_analysis=True, return_data=True)

    # First segment = the CURRENT config watchlist only. Map every archive company
    # to its canonical watchlist name (via name + aliases, ignoring case/spacing) and
    # merge signals — so renamed/variant rows like "FigureAI"/"Figure AI" or
    # "Crusoe"/"Crusoe Energy" collapse into one company instead of splitting or
    # being dropped. Unmatched rows (stale/off-watchlist/public names) are dropped.
    from config import WATCHLIST

    def _norm(s):
        return "".join(ch for ch in (s or "").lower() if ch.isalnum())

    canon, sector_of = {}, {}
    for w in WATCHLIST:
        canon[_norm(w["name"])] = w["name"]
        sector_of[w["name"]] = w.get("sector", "")
        for a in w.get("aliases", []):
            canon.setdefault(_norm(a), w["name"])

    merged, dropped = {}, 0
    for r in recaps:
        key = canon.get(_norm(r.get("company")))
        if not key:
            dropped += 1
            continue
        m = merged.get(key)
        if m is None:
            m = {"company": key, "sector": sector_of.get(key, r.get("sector", "")),
                 "signals": [], "signal_count": 0, "watchlist": True, "connection": ""}
            merged[key] = m
        m["signals"].extend(r.get("signals", []))
    for m in merged.values():
        m["signal_count"] = len(m["signals"])
    recaps = list(merged.values())
    print(f"  Watchlist: {len(recaps)} companies after canonicalizing "
          f"(dropped {dropped} off-watchlist archive group(s))")

    from weekly_layout import assemble, flatten_signals, rank_and_split

    # Second segment: notable NON-watchlist PRIVATE names (public names excluded).
    print("\nGathering non-watchlist private-market movers...")
    try:
        movers = gather_private_movers(days=days)
    except Exception as e:
        print(f"  [ERROR] private movers failed: {e}")
        movers = []
    recaps = recaps + movers

    print("Running the semantic dedupe pass...")
    llm_dedupe(recaps, candidate_n=40)
    print("Running the per-company semantic dedupe pass...")
    llm_dedupe_per_company(recaps)

    print("Filtering out events already covered in prior weeks...")
    prior = read_prior_signals(window_days=days, lookback_days=60)
    filter_already_covered(recaps, prior)

    print("Generating private-public connection notes...")
    generate_connections(recaps)

    top, _ = rank_and_split(flatten_signals(recaps), top_n=10, radar_n=0)
    print("Synthesising the week in brief...")
    brief = generate_week_in_brief(industry_news, top)
    industry_news = {"week_in_brief": brief}

    print("Drafting key events for next week...")
    key_events = draft_key_events(recaps, industry_news, days=days)

    data = assemble(recaps, industry_news, key_events=[], days=days, issue=issue)
    _write_weekly_json(data, key_events)

    print(f"\n{'=' * 60}")
    print("  Prepare complete — review weekly/key_events_next_week.json before Sunday.")
    print(f"{'=' * 60}")
    tracker = _get_tracker()
    if tracker.call_count:
        tracker.print_summary()


def run_weekly_send(archive_url: str = None):
    """SUNDAY: render from saved JSON (+ any edits to key events), email, publish."""
    print("\n=== WEEKLY SEND ===")
    data, key_events = _read_weekly_json()
    if not data:
        print(f"[ERROR] {WEEKLY_DATA_PATH} not found. Run --weekly-prepare first.")
        return

    from weekly_layout import normalise_key_events, render_weekly
    data["key_events"] = normalise_key_events(key_events)
    data["generated_at"] = datetime.utcnow().strftime("%b %d, %Y \u00b7 %H:%M UTC")

    html = render_weekly(data, archive_url=archive_url if archive_url is not None else ARCHIVE_URL)
    _publish_html(html)

    week_label = data.get("week_label", "")
    send_weekly_email(html, week_label)
    print(f"\n{'=' * 60}\n  Weekly recap sent for: {week_label}\n{'=' * 60}")


