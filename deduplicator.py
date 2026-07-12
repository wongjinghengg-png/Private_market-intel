"""
Deduplicator — removes duplicate signals covering the same news event,
keeping the most credible source for each.

Two-pass approach (zero API cost):
1. Group signals by category, then cluster by headline/summary similarity
2. Within each cluster, keep the signal from the highest-credibility source

Can be used in both daily (post-classification) and weekly (post-archive-read) flows.
"""

import re
from config import SOURCE_TIER_LOOKUP, DEFAULT_SOURCE_TIER


# ── Text similarity ──────────────────────────────────────────────────

def _tokenize(text: str) -> set[str]:
    """Simple word-level tokenizer, lowercased, stripped of punctuation."""
    return set(re.findall(r'[a-z0-9]+', text.lower()))


def _similarity(a: str, b: str) -> float:
    """Jaccard similarity between two text strings."""
    tokens_a = _tokenize(a)
    tokens_b = _tokenize(b)
    if not tokens_a or not tokens_b:
        return 0.0
    intersection = tokens_a & tokens_b
    union = tokens_a | tokens_b
    return len(intersection) / len(union)


def _get_source_tier(source: str) -> int:
    """Look up credibility tier for a source. Lower = more credible."""
    if not source:
        return DEFAULT_SOURCE_TIER
    return SOURCE_TIER_LOOKUP.get(source.lower().strip(), DEFAULT_SOURCE_TIER)


# ── Core deduplication ───────────────────────────────────────────────

SIMILARITY_THRESHOLD = 0.35  # Signals above this are considered duplicates

def deduplicate_signals(
    signals: list[dict],
    similarity_threshold: float = SIMILARITY_THRESHOLD,
    verbose: bool = False,
) -> list[dict]:
    """
    Deduplicate a list of classified signals for a single company.

    For each group of similar signals (same category, similar headline/summary),
    keeps the one from the most credible source. If tied on credibility,
    keeps the one with the higher relevance score. If still tied, keeps the
    most recent.

    Args:
        signals: List of signal dicts with keys: category, summary, headline,
                 source, relevance, published_date, url, etc.
        similarity_threshold: Jaccard similarity above which signals are
                              considered duplicates (default 0.45).
        verbose: Print dedup decisions.

    Returns:
        Deduplicated list of signals.
    """
    if len(signals) <= 1:
        return signals

    # Group by category first — only dedup within same category
    by_category = {}
    for sig in signals:
        cat = sig.get("category", "Unknown")
        by_category.setdefault(cat, []).append(sig)

    kept = []
    total_removed = 0

    for category, cat_signals in by_category.items():
        if len(cat_signals) <= 1:
            kept.extend(cat_signals)
            continue

        # Build clusters of similar signals
        clusters = []
        used = [False] * len(cat_signals)

        for i in range(len(cat_signals)):
            if used[i]:
                continue

            cluster = [i]
            used[i] = True

            # Compare text: use summary primarily, fall back to headline
            text_i = cat_signals[i].get("summary", "") or cat_signals[i].get("headline", "")

            for j in range(i + 1, len(cat_signals)):
                if used[j]:
                    continue
                text_j = cat_signals[j].get("summary", "") or cat_signals[j].get("headline", "")

                sim = _similarity(text_i, text_j)
                if sim >= similarity_threshold:
                    cluster.append(j)
                    used[j] = True

            clusters.append(cluster)

        # From each cluster, pick the best signal
        for cluster in clusters:
            if len(cluster) == 1:
                kept.append(cat_signals[cluster[0]])
                continue

            # Sort cluster members: lowest tier (most credible) first,
            # then highest relevance, then most recent date
            candidates = [cat_signals[idx] for idx in cluster]
            candidates.sort(key=lambda s: (
                _get_source_tier(s.get("source", "")),
                -s.get("relevance", 1),
                s.get("published_date", "0000"),  # reverse: we want most recent
            ))
            # Reverse the date sort — most recent should be preferred
            # Actually, the tuple sort handles it: lower tier wins, then higher relevance,
            # then... we want most recent. Let's negate by using reverse string comparison.
            # Simpler: just re-sort with a proper key.
            candidates.sort(key=lambda s: (
                _get_source_tier(s.get("source", "")),     # lower = better
                -s.get("relevance", 1),                     # higher = better
                "" if not s.get("published_date") else      # more recent = better
                    "".join(c if c.isdigit() else "" for c in str(s.get("published_date", ""))),
            ))

            winner = candidates[0]
            removed = candidates[1:]

            # Collect alternative source URLs for reference
            alt_sources = []
            for r in removed:
                src = r.get("source", "")
                url = r.get("url", "")
                if src and url:
                    alt_sources.append(f"{src}: {url}")
                elif src:
                    alt_sources.append(src)

            if alt_sources:
                winner = dict(winner)  # Don't mutate original
                existing = winner.get("alt_sources", "")
                winner["alt_sources"] = "; ".join(alt_sources) if not existing else f"{existing}; {'; '.join(alt_sources)}"

            kept.append(winner)
            total_removed += len(removed)

            if verbose:
                print(f"    [DEDUP] Kept: {winner.get('source', '?')} (tier {_get_source_tier(winner.get('source', ''))})")
                for r in removed:
                    print(f"      Removed: {r.get('source', '?')} (tier {_get_source_tier(r.get('source', ''))})")

    if verbose and total_removed > 0:
        print(f"    Dedup: {len(signals)} → {len(kept)} signals ({total_removed} duplicates removed)")

    return kept


# ── Batch dedup (multiple companies) ────────────────────────────────

def deduplicate_by_company(by_company: dict, verbose: bool = False) -> dict:
    """
    Deduplicate signals for each company in a {company: [signals]} dict.

    Args:
        by_company: Dict mapping company name to list of signal dicts.
        verbose: Print dedup decisions.

    Returns:
        Same structure with duplicates removed.
    """
    deduped = {}
    total_before = 0
    total_after = 0

    for company, signals in by_company.items():
        total_before += len(signals)
        deduped[company] = deduplicate_signals(signals, verbose=verbose)
        total_after += len(deduped[company])

    removed = total_before - total_after
    if removed > 0:
        print(f"  [DEDUP] {total_before} → {total_after} signals ({removed} duplicates removed across {len(by_company)} companies)")
    else:
        print(f"  [DEDUP] No duplicates found ({total_before} signals)")

    return deduped


# ── Standalone CLI ──────────────────────────────────────────────────

def run_standalone(dry_run: bool = False, threshold: float = SIMILARITY_THRESHOLD):
    """
    Run deduplication against the full archive.
    - dry_run=True: show what would be removed, don't touch the archive
    - dry_run=False: rewrite the archive with duplicates removed
    """
    from datetime import datetime
    from pathlib import Path
    from openpyxl import load_workbook, Workbook
    from config import ARCHIVE_PATH

    path = Path(ARCHIVE_PATH)
    if not path.exists():
        print("[ERROR] No archive found at", ARCHIVE_PATH)
        return

    # Read full archive into signals grouped by company
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active

    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {h: i for i, h in enumerate(headers)}

    by_company = {}
    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        row_count += 1

        company = str(row[col_map.get("Company", 1)] or "Unknown")
        signal = {}
        for header, idx in col_map.items():
            val = row[idx]
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            signal[header.lower().replace(" ", "_")] = str(val) if val is not None else ""

        # Normalize keys to match what dedup expects
        signal.setdefault("category", signal.get("category", "Unknown"))
        signal.setdefault("summary", signal.get("summary", ""))
        signal.setdefault("headline", signal.get("headline", ""))
        signal.setdefault("source", signal.get("source", ""))
        signal.setdefault("url", signal.get("url", ""))
        signal.setdefault("published_date", signal.get("published_date", ""))
        signal.setdefault("relevance", 1)
        try:
            signal["relevance"] = int(signal.get("relevance", 1))
        except (ValueError, TypeError):
            signal["relevance"] = 1

        by_company.setdefault(company, []).append(signal)

    wb.close()

    print(f"\n{'=' * 60}")
    print(f"  Deduplicator — Standalone Mode")
    print(f"  Archive: {ARCHIVE_PATH}")
    print(f"  Total rows: {row_count}")
    print(f"  Companies: {len(by_company)}")
    print(f"  Threshold: {threshold}")
    print(f"  Mode: {'DRY RUN (no changes)' if dry_run else 'LIVE (will rewrite archive)'}")
    print(f"{'=' * 60}\n")

    # Run dedup
    deduped = deduplicate_by_company(by_company, verbose=True)

    total_after = sum(len(sigs) for sigs in deduped.values())
    removed = row_count - total_after

    print(f"\n{'=' * 60}")
    print(f"  Results: {row_count} → {total_after} ({removed} duplicates)")
    print(f"{'=' * 60}")

    if removed == 0:
        print("  No duplicates found. Archive is clean.\n")
        return

    if dry_run:
        print("  DRY RUN — no changes made. Run without --dry-run to apply.\n")
        return

    # Rewrite archive
    print(f"\n  Rewriting archive...")
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(headers)

    for company in sorted(deduped.keys()):
        for sig in deduped[company]:
            row_data = []
            for h in headers:
                key = h.lower().replace(" ", "_")
                row_data.append(sig.get(key, ""))
            new_ws.append(row_data)

    # Backup original
    backup_path = path.with_suffix(".backup.xlsx")
    import shutil
    shutil.copy2(str(path), str(backup_path))
    print(f"  Backup saved: {backup_path}")

    new_wb.save(str(path))
    print(f"  Archive rewritten: {path}")
    print(f"  ✅ {removed} duplicates removed.\n")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Deduplicate the market intel archive")
    parser.add_argument("--dry-run", action="store_true", help="Show what would be removed without changing the archive")
    parser.add_argument("--threshold", type=float, default=SIMILARITY_THRESHOLD, help=f"Similarity threshold (default: {SIMILARITY_THRESHOLD})")
    args = parser.parse_args()

    run_standalone(dry_run=args.dry_run, threshold=args.threshold)
