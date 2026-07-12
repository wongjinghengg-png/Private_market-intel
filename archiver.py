"""
Archiver module — appends classified items to a running Excel archive.
Creates the file if it doesn't exist. Each run adds new rows.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import ARCHIVE_PATH, MIN_RELEVANCE, FALLBACK_START_DATE

HEADERS = [
    "Date Scraped",
    "Company",
    "Sector",
    "Category",
    "Headline",
    "Summary",
    "Relevance",
    "Source",
    "Published Date",
    "URL",
]

HEADER_FILL = PatternFill("solid", fgColor="0F172A")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="E2E8F0"),
)

COL_WIDTHS = [14, 18, 14, 22, 50, 60, 10, 20, 14, 50]


def get_last_run_dates(watchlist: list[dict]) -> dict[str, str]:
    """Read the archive and return the latest 'Date Scraped' per company.
    New companies not yet in the archive get FALLBACK_START_DATE.
    If no archive exists, all companies start from FALLBACK_START_DATE."""
    path = Path(ARCHIVE_PATH)
    company_dates: dict[str, str] = {}

    if not path.exists():
        print(f"[INFO] No archive found. All companies start from {FALLBACK_START_DATE}")
        return {co["name"]: FALLBACK_START_DATE for co in watchlist}

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        # Column A = Date Scraped, Column B = Company
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
            date_val, company_val = row[0], row[1]
            if date_val is None or company_val is None:
                continue
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val).strip()
            company_name = str(company_val).strip()
            if company_name not in company_dates or date_str > company_dates[company_name]:
                company_dates[company_name] = date_str
        wb.close()
    except Exception as e:
        print(f"[WARN] Could not read archive: {e}. Using {FALLBACK_START_DATE} for all.")
        return {co["name"]: FALLBACK_START_DATE for co in watchlist}

    # Fill in any watchlist companies not yet in the archive
    result = {}
    for co in watchlist:
        name = co["name"]
        if name in company_dates:
            result[name] = company_dates[name]
            print(f"  {name}: last archived {company_dates[name]}")
        else:
            result[name] = FALLBACK_START_DATE
            print(f"  {name}: NEW — backfilling from {FALLBACK_START_DATE}")

    return result


def archive_items(items: list[dict], watchlist: list[dict]) -> str:
    """Append items to the archive Excel file. Returns the file path."""
    path = Path(ARCHIVE_PATH)
    path.parent.mkdir(parents=True, exist_ok=True)

    # Build sector lookup
    sector_map = {co["name"]: co.get("sector", "") for co in watchlist}

    if path.exists():
        wb = load_workbook(str(path))
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Intelligence Archive"
        _write_headers(ws)

    # Build set of existing URLs to prevent duplicates across runs
    existing_urls = set()
    url_col = 10  # Column J = URL
    for row in ws.iter_rows(min_row=2, min_col=url_col, max_col=url_col, values_only=True):
        if row[0]:
            existing_urls.add(row[0].split("?")[0].rstrip("/").lower())

    date_str = datetime.now().strftime("%Y-%m-%d")
    added = 0
    skipped_dupes = 0

    for item in items:
        if item.get("category") in ("Irrelevant", "Unknown"):
            continue
        if item.get("relevance", 0) < MIN_RELEVANCE:
            continue

        # Deduplicate by URL
        url_key = item.get("url", "").split("?")[0].rstrip("/").lower()
        if url_key in existing_urls:
            skipped_dupes += 1
            continue
        existing_urls.add(url_key)

        ws.append([
            date_str,
            item.get("company", ""),
            sector_map.get(item.get("company", ""), ""),
            item.get("category", ""),
            item.get("title", ""),
            item.get("summary", ""),
            item.get("relevance", 1),
            item.get("source", ""),
            item.get("published_date", ""),
            item.get("url", ""),
        ])

        # Apply data formatting to new row
        row_num = ws.max_row
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col in [5, 6, 10]))

        # Color-code relevance
        rel_cell = ws.cell(row=row_num, column=7)
        rel_val = item.get("relevance", 1)
        if rel_val >= 4:
            rel_cell.fill = PatternFill("solid", fgColor="DCFCE7")
            rel_cell.font = Font(name="Arial", size=10, bold=True, color="166534")
        elif rel_val >= 3:
            rel_cell.fill = PatternFill("solid", fgColor="FEF9C3")

        added += 1

    # Apply auto-filter to full range
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(str(path))
    print(f"[OK] Archive updated: {path} ({ws.max_row - 1} total records, +{added} new, {skipped_dupes} duplicates skipped)")
    return str(path)


def _write_headers(ws):
    """Write formatted header row."""
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]
